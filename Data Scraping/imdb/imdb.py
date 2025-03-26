from bs4 import BeautifulSoup
import requests
import openpyxl

excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = 'Top Rated Movies'
print(excel.sheetnames)
sheet.append(['Movie Rank', 'Movie Title', 'Year of Release', 'IMDB Rating', 'Movie Link'])

try:
    url = requests.get("https://www.imdb.com/chart/top/?ref_=nv_mv_250")
    url.raise_for_status()

    soup = BeautifulSoup(url.text, 'html.parser')
    movies = soup.find('tbody', class_="lister-list").find_all('tr')
    for movie in movies:
        name = movie.find('td', class_="titleColumn").a.string
        rank = movie.find('td', class_="titleColumn").get_text(strip=True).split('.')[0]
        year = movie.find('td', class_="titleColumn").span.text.strip('()')
        rating = movie.find('td', class_="ratingColumn imdbRating").strong.text
        link = movie.find('td', class_="titleColumn").a['href']
        comp_link = (f"https://www.imdb.com/{link}")
        print(rank, name, year, rating, comp_link)
        sheet.append([rank, name, year, rating, comp_link])

except Exception as e:
    print(e)

excel.save('IMDB Move Ratings.xlsx')
